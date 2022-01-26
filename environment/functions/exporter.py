import os
import io
import csv
from openpyxl import Workbook, utils
from openpyxl.styles import Font
from functions.misc import SERVE_DIR, get_time
from typing import Any, Union


# This function is an example of how to export db results to file.
# there are easier/better ways to do this, such as the Pandas library.
class ExportFile():
    def __init__(self, result_set: list) -> None:
        self.result_set = result_set

    def as_text(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value)

    def to_csv(self, as_strio: bool = False) -> Union[io.StringIO, str]:
        csv_columns = self.result_set[0]
        if as_strio:
            file_path = io.StringIO()
        else:
            file_path = os.path.join(SERVE_DIR, f'{get_time()}.csv')
        result_list = []
        for row in self.result_set:
            result_list.append(dict(zip(csv_columns, row)))
        if as_strio:
            writer = csv.DictWriter(
                file_path,
                fieldnames=csv_columns,
                lineterminator='\n'
            )
            for data in result_list:
                writer.writerow(data)
            file_path.seek(0)
        else:
            with open(file_path, 'w') as csvfile:
                writer = csv.DictWriter(
                    csvfile,
                    fieldnames=csv_columns,
                    lineterminator='\n'
                )
                for data in result_list:
                    writer.writerow(data)
        if as_strio:
            return file_path
        else:
            return os.path.basename(file_path)

    def to_excel(self) -> str:
        csv_data = self.to_csv(as_strio = True)

        file_path = os.path.join(SERVE_DIR, f'{get_time()}.xlsx')
        wb = Workbook()
        ws = wb.active
        for row in csv.reader(csv_data):
            print(row)
            ws.append(row)
        # Auto-fit columns
        for column_cells in ws.columns:
            length = max(len(self.as_text(cell.value)) for cell in column_cells)
            ws.column_dimensions[
                utils.get_column_letter(column_cells[0].column)
            ].width = (length * 1.2)
        # Bold, blue header
        for cell in ws["1:1"]:
            cell.font = Font(color='0539f7', bold=True)
        wb.save(file_path)
        return os.path.basename(file_path)
